import openpyxl
book =openpyxl.load_workbook("C:\\Users\\ZZ366EN\\OneDrive - EY\\Documents\\Pythonpackage.xlsx")
sheet=book.active
cell=sheet.cell(row=1,column=2)
print(cell.value)
sheet.cell(row=2,column=2).value="abyes" # to write anything in the sheet
print(sheet.max_row) #it tells how many total rows in sheet have same for column